import argparse
import csv
import hashlib
import json
import os
import re
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from uuid import UUID

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import execute_values

DEFAULT_LIMIT = 5000
MAX_LIMIT = 50000

CATEGORY_RULES = {
    "生活用品": ["垃圾袋", "抽纸", "纸巾", "清洁", "洗衣", "湿巾", "厨房", "锅", "厨具", "收纳", "保鲜", "拖把"],
    "家居家具": ["沙发", "桌", "椅", "床", "床垫", "衣柜", "窗帘", "家纺", "枕头"],
    "个护美妆": ["洗发", "护发", "沐浴", "牙膏", "护肤", "面膜", "口红", "香水", "防晒"],
    "服饰": ["T恤", "卫衣", "外套", "裤", "裙", "衬衫", "毛衣", "羽绒"],
    "鞋靴": ["运动鞋", "板鞋", "皮鞋", "凉鞋", "靴"],
    "箱包配饰": ["背包", "双肩包", "手提包", "行李箱", "钱包", "帽", "围巾", "腰带"],
    "电子数码": ["手机", "耳机", "笔记本", "电脑", "充电器", "平板", "相机", "键盘", "鼠标"],
    "食品生鲜": ["零食", "饮料", "牛奶", "咖啡", "坚果", "水果", "米", "面", "粮油"],
    "母婴玩具": ["婴儿", "奶瓶", "纸尿裤", "玩具", "积木", "童装"],
    "宠物用品": ["猫", "狗", "宠物", "猫粮", "狗粮", "猫砂"],
}

MAJOR_CATEGORY_RULES = {
    "服饰鞋包": ["女装", "男装", "童装", "t恤", "卫衣", "外套", "衬衫", "裤", "裙", "鞋", "靴", "包", "背包", "箱包"],
    "数码家电": ["手机", "耳机", "笔记本", "电脑", "平板", "键盘", "鼠标", "显示器", "相机", "手表", "手环", "充电器", "电器", "家电"],
    "家居生活": ["家居", "日用", "清洁", "抽纸", "纸巾", "厨房", "锅", "收纳", "床品", "枕头", "拖把"],
    "个护母婴": ["母婴", "奶粉", "纸尿裤", "湿巾", "护肤", "洗护", "面膜", "婴儿", "宝宝"],
    "食品饮料": ["零食", "饮料", "牛奶", "咖啡", "坚果", "水果", "米", "面", "粮油"],
    "宠物用品": ["宠物", "猫粮", "狗粮", "猫砂", "猫", "狗"],
}

HIDDEN_TAGS = {"crawler", "escuelajs", "jsonfile", "xlsx", "csv", "精选商品", "精选"}

PRICE_PATTERN = re.compile(r"-?\d+(?:\.\d+)?")
COUNT_PATTERN = re.compile(r"\d+(?:\.\d+)?")


def get_db_config() -> dict:
    return {
        "host": os.getenv("DB_HOST", "localhost"),
        "port": int(os.getenv("DB_PORT", "5432")),
        "dbname": os.getenv("DB_NAME", "ecommerce"),
        "user": os.getenv("DB_USER", "postgres"),
        "password": os.getenv("DB_PASSWORD", "123456"),
    }


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKC", value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace("￥", "¥")
    return text


def clip_text(value: str | None, max_len: int) -> str:
    text = normalize_text(value)
    if len(text) <= max_len:
        return text
    return text[:max_len]


def abbreviate(text: str, max_len: int) -> str:
    clean = normalize_text(text)
    if len(clean) <= max_len:
        return clean
    return clean[:max_len].rstrip() + "..."


def to_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0.00")
    text = normalize_text(str(value)).lower().replace(",", "")
    if not text:
        return Decimal("0.00")
    match = PRICE_PATTERN.search(text)
    if not match:
        return Decimal("0.00")
    try:
        return Decimal(match.group(0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def to_int(value) -> int:
    if value is None:
        return 0
    text = normalize_text(str(value)).replace(",", "")
    if not text:
        return 0
    try:
        if "万" in text:
            base = COUNT_PATTERN.search(text)
            return int(float(base.group(0)) * 10000) if base else 0
        if "千" in text:
            base = COUNT_PATTERN.search(text)
            return int(float(base.group(0)) * 1000) if base else 0
        base = COUNT_PATTERN.search(text)
        return int(float(base.group(0))) if base else 0
    except Exception:
        return 0


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    text = normalize_text(str(value)).lower()
    return text in {"1", "true", "yes", "y", "是", "支持", "有"}


def parse_update_time(value, fallback: datetime) -> datetime:
    if not value:
        return fallback
    text = normalize_text(str(value))
    if not text:
        return fallback
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return fallback


def infer_category(*texts: str) -> str:
    content = " ".join(normalize_text(text) for text in texts if text)
    if not content:
        return "其他"
    for category, keywords in CATEGORY_RULES.items():
        if any(keyword in content for keyword in keywords):
            return category
    return "其他"


def infer_major_category(*texts: str) -> str:
    content = " ".join(normalize_text(text).lower() for text in texts if text)
    if not content:
        return "其他"
    for category, keywords in MAJOR_CATEGORY_RULES.items():
        if any(keyword in content for keyword in keywords):
            return category
    return infer_category(*texts)


def build_clean_tags(*values: str) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        for tag in re.split(r"[,，]", normalize_text(value)):
            clean = normalize_text(tag)
            key = clean.lower()
            if not clean or key in HIDDEN_TAGS or key in seen:
                continue
            result.append(clean)
            seen.add(key)
    return ",".join(result)


def normalize_image(item: dict) -> str:
    direct = normalize_text(
        item.get("image")
        or item.get("image_url")
        or item.get("thumbnail")
        or item.get("商品图片")
        or item.get("图片地址")
        or ""
    )
    if direct:
        return direct
    images = item.get("images")
    if isinstance(images, list):
        for image in images:
            normalized = normalize_text(str(image))
            if normalized:
                return normalized
    return ""


def extract_specs(text: str, fallback_category: str) -> str:
    clean = normalize_text(text)
    match = re.search(r"规格[:：]\s*([^,，。;；]{2,80})", clean, flags=re.IGNORECASE)
    if match:
        return f"规格：{match.group(1).strip()}"
    if fallback_category:
        return f"规格：{fallback_category}标准款"
    return "规格：标准款"


def build_policy() -> str:
    return "官方正品，7天无理由，48小时内发货"


def deterministic_uuid(scope: str, source_key: str) -> UUID:
    payload = f"{scope}:{source_key}".encode("utf-8")
    digest = hashlib.md5(payload).hexdigest()
    return UUID(digest)


def load_local_csv(path: str) -> list[dict]:
    candidate = Path(path).expanduser().resolve()
    if not candidate.exists():
        raise FileNotFoundError(f"CSV file not found: {candidate}")

    with candidate.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for row in reader:
            if not isinstance(row, dict):
                continue
            normalized = {str(k).strip(): ("" if v is None else str(v).strip()) for k, v in row.items() if k is not None}
            if normalized:
                rows.append(normalized)
        return rows


def load_local_xlsx(path: str) -> list[dict]:
    candidate = Path(path).expanduser().resolve()
    if not candidate.exists():
        raise FileNotFoundError(f"XLSX file not found: {candidate}")

    workbook = load_workbook(candidate, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if not headers_row:
        return []

    headers = [normalize_text(str(item) if item is not None else "") for item in headers_row]
    normalized_rows: list[dict] = []
    for row in rows:
        mapped: dict = {}
        for index, value in enumerate(row):
            key = headers[index] if index < len(headers) else f"col_{index}"
            if not key:
                continue
            mapped[key] = value
        if mapped:
            mapped["_sheet_name"] = sheet.title
            normalized_rows.append(mapped)
    return normalized_rows


def load_local_data(path: str) -> tuple[list[dict], str]:
    candidate = Path(path).expanduser().resolve()
    suffix = candidate.suffix.lower()
    if suffix == ".csv":
        return load_local_csv(path), "CSV"
    if suffix == ".xlsx":
        return load_local_xlsx(path), "XLSX"
    raise ValueError("Only .csv and .xlsx are supported")


def normalize_product_record(source: str, item: dict, now: datetime, data_source: str) -> dict:
    title = normalize_text(
        item.get("title")
        or item.get("name")
        or item.get("product_name")
        or item.get("产品名称")
        or item.get("商品名称")
        or item.get("商品标题")
    )
    raw_keyword = normalize_text(item.get("关键词") or item.get("keyword") or "")
    category = normalize_text(
        item.get("major_category")
        or item.get("category")
        or item.get("category_path")
        or item.get("catalog")
        or item.get("关键词")
        or item.get("商品分类")
        or item.get("类目")
    )
    raw_desc = normalize_text(
        item.get("description")
        or item.get("desc")
        or item.get("selling_points")
        or item.get("商品描述")
        or item.get("评论Top5(JSON)")
    )
    source_product_id = normalize_text(
        str(
            item.get("id")
            or item.get("product_id")
            or item.get("sku")
            or item.get("产品ID")
            or item.get("商品ID")
            or item.get("商品链接")
            or ""
        )
    )

    if len(source_product_id) > 120:
        source_product_id = hashlib.md5(source_product_id.encode("utf-8")).hexdigest()[:32]
    if not source_product_id:
        fingerprint = "|".join(
            [
                title,
                category,
                normalize_text(str(item.get("price") or item.get("产品价格") or item.get("商品价格") or "")),
                normalize_text(str(item.get("店铺名称") or "")),
                normalize_text(str(item.get("图片地址") or item.get("商品图片") or "")),
                normalize_text(str(item.get("发货_发货时间") or "")),
                normalize_text(raw_keyword),
                normalize_text(str(item.get("_sheet_name") or "")),
            ]
        )
        source_product_id = hashlib.md5(fingerprint.encode("utf-8")).hexdigest()[:16]

    specs = normalize_text(item.get("specs") or "") or extract_specs(raw_desc, category)
    selling_points = abbreviate(item.get("selling_points") or raw_desc, 150)
    policy = normalize_text(item.get("policy")) or build_policy()
    price = to_decimal(
        item.get("price")
        or item.get("产品价格")
        or item.get("current_price")
        or item.get("sale_price")
        or item.get("商品价格")
    )
    image_url = normalize_image(item)
    updated_at = parse_update_time(item.get("update_time"), now)

    paid_count = to_int(item.get("付款人数") or item.get("已付款人数") or item.get("paid_count"))
    review_count = to_int(item.get("已评价数") or item.get("review_count"))
    sales_volume = to_int(item.get("商品销量") or item.get("sales_volume"))
    if sales_volume <= 0 and paid_count > 0:
        sales_volume = paid_count
    positive_rate = to_decimal(item.get("好评率(页面展示)") or item.get("positive_rate"))
    total_reviews = to_int(item.get("总评数") or item.get("total_reviews"))
    shipping_time = clip_text(item.get("发货_发货时间") or item.get("shipping_time"), 128)
    support_7day_return = parse_bool(item.get("保障_7天无理由退货") or item.get("support_7day_return"))
    support_fast_refund = parse_bool(item.get("保障_极速退款") or item.get("support_fast_refund"))
    payment_method = clip_text(item.get("支付_支持方式") or item.get("payment_method"), 128)
    store_name = clip_text(item.get("店铺名称") or item.get("store_name"), 255)
    store_url = clip_text(item.get("店铺链接") or item.get("store_url"), 512)
    top_reviews_json = normalize_text(item.get("评论Top5(JSON)") or item.get("top_reviews_json"))

    if not category:
        category = infer_major_category(raw_keyword, title, raw_desc, specs, selling_points, normalize_text(item.get("_sheet_name") or ""))
    else:
        category = infer_major_category(category, raw_keyword, title)

    if not raw_desc:
        raw_desc = f"{title}，适用于{raw_keyword or category}场景。"

    if not selling_points:
        selling_points = abbreviate(f"{category}优选；结合销量与价格进行筛选", 150)

    product_id = deterministic_uuid(source, source_product_id)
    raw_tags = item.get("tags")
    if isinstance(raw_tags, list):
        tags_value = ",".join(normalize_text(str(tag)) for tag in raw_tags if normalize_text(str(tag)))
    else:
        tags_value = normalize_text(raw_tags)
    tags = build_clean_tags(tags_value, raw_keyword, category)

    return {
        "id": product_id,
        "name": clip_text(title, 255),
        "description": abbreviate(raw_desc, 240),
        "price": price,
        "image_url": clip_text(image_url, 512),
        "tags": clip_text(tags, 512),
        "category": clip_text(category, 128),
        "specs": clip_text(specs, 512),
        "selling_points": selling_points,
        "policy": clip_text(policy, 512),
        "source_product_id": clip_text(source_product_id, 128),
        "data_source": clip_text(data_source, 32),
        "paid_count": paid_count,
        "review_count": review_count,
        "sales_volume": sales_volume,
        "positive_rate": positive_rate,
        "total_reviews": total_reviews,
        "shipping_time": shipping_time,
        "support_7day_return": support_7day_return,
        "support_fast_refund": support_fast_refund,
        "payment_method": payment_method,
        "store_name": store_name,
        "store_url": store_url,
        "top_reviews_json": top_reviews_json,
        "created_at": now,
        "updated_at": updated_at,
    }


def load_products(input_file: str, limit: int) -> list[dict]:
    now = datetime.now(timezone.utc)
    items, source = load_local_data(input_file)
    products = [
        normalize_product_record(source.lower(), item, now, data_source=source)
        for item in items
        if isinstance(item, dict)
    ]
    return products[:limit]


def deduplicate_products(products: list[dict]) -> list[dict]:
    by_key: dict[str, dict] = {}
    for product in products:
        name = normalize_text(str(product.get("name") or "")).lower()
        store = normalize_text(str(product.get("store_name") or "")).lower()
        category = normalize_text(str(product.get("category") or "")).lower()
        price = str(product.get("price") or "0")
        source_product_id = normalize_text(str(product.get("source_product_id") or "")).lower()
        key = "|".join([name, store, category, price, source_product_id])
        if key.strip("|") == "":
            continue
        existing = by_key.get(key)
        if not existing:
            by_key[key] = product
            continue

        existing_score = existing.get("sales_volume") or 0
        current_score = product.get("sales_volume") or 0
        existing_time = existing.get("updated_at") or datetime.min.replace(tzinfo=timezone.utc)
        current_time = product.get("updated_at") or datetime.min.replace(tzinfo=timezone.utc)
        if current_score > existing_score or current_time >= existing_time:
            by_key[key] = product
    return list(by_key.values())


def build_vector_row(product: dict) -> tuple[str, str, int, str, str, str, datetime]:
    product_id = str(product["id"])
    chunk_index = 0
    updated_at = product["updated_at"]
    chunk_text = "\n".join([
        f"product_id={product_id}",
        f"update_time={updated_at.isoformat() if updated_at else ''}",
        f"title={normalize_text(product.get('name'))}",
        f"category={normalize_text(product.get('category'))}",
        f"price={product.get('price')}元",
        f"specs={normalize_text(product.get('specs'))}",
        f"selling_points={normalize_text(product.get('selling_points'))}",
        f"policy={normalize_text(product.get('policy'))}",
        f"sales_volume={product.get('sales_volume')}",
        f"positive_rate={product.get('positive_rate')}",
        f"store_name={normalize_text(product.get('store_name'))}",
        f"source={normalize_text(product.get('data_source') or 'CSV')}",
    ])
    metadata = {
        "product_id": product_id,
        "update_time": updated_at.isoformat() if updated_at else "",
        "category": normalize_text(product.get("category")),
        "source": normalize_text(product.get("data_source") or "CSV"),
        "source_product_id": normalize_text(product.get("source_product_id")),
        "sales_volume": product.get("sales_volume") or 0,
        "positive_rate": str(product.get("positive_rate") or "0"),
        "store_name": normalize_text(product.get("store_name")),
    }
    vector_id = str(deterministic_uuid("vector", f"{product_id}:{chunk_index}"))
    return (
        vector_id,
        product_id,
        chunk_index,
        chunk_text,
        json.dumps(metadata, ensure_ascii=False),
        normalize_text(product.get("data_source") or "CSV"),
        updated_at,
    )


def ensure_schema(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
              id UUID PRIMARY KEY,
              name VARCHAR(255) NOT NULL,
              description TEXT,
              price NUMERIC(12, 2) NOT NULL,
              image_url VARCHAR(512),
              tags VARCHAR(512),
              category VARCHAR(128),
              specs VARCHAR(512),
              selling_points TEXT,
              policy VARCHAR(512),
              paid_count INT,
              review_count INT,
              sales_volume INT,
              positive_rate NUMERIC(8, 4),
              total_reviews INT,
              shipping_time VARCHAR(128),
              support_7day_return BOOLEAN,
              support_fast_refund BOOLEAN,
              payment_method VARCHAR(128),
              store_name VARCHAR(255),
              store_url VARCHAR(512),
              top_reviews_json TEXT,
              source_product_id VARCHAR(128),
              data_source VARCHAR(32),
              created_at TIMESTAMP NOT NULL,
              updated_at TIMESTAMP
            )
            """
        )
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS category VARCHAR(128)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS specs VARCHAR(512)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS selling_points TEXT")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS policy VARCHAR(512)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS source_product_id VARCHAR(128)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS data_source VARCHAR(32)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS paid_count INT")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS review_count INT")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS sales_volume INT")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS positive_rate NUMERIC(8, 4)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS total_reviews INT")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS shipping_time VARCHAR(128)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS support_7day_return BOOLEAN")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS support_fast_refund BOOLEAN")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS payment_method VARCHAR(128)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS store_name VARCHAR(255)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS store_url VARCHAR(512)")
        cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS top_reviews_json TEXT")
        cur.execute("UPDATE products SET updated_at = created_at WHERE updated_at IS NULL")

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS product_vectors (
              id UUID PRIMARY KEY,
              product_id UUID NOT NULL,
              chunk_index INT NOT NULL,
              chunk_text TEXT NOT NULL,
              metadata_json TEXT NOT NULL,
              source VARCHAR(32) NOT NULL,
              updated_at TIMESTAMP NOT NULL
            )
            """
        )
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_product_vectors_product_chunk ON product_vectors(product_id, chunk_index)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_product_vectors_source_updated ON product_vectors(source, updated_at)")
    conn.commit()


def upsert_products(conn, products: list[dict]):
    rows = [
        (
            str(item["id"]),
            item["name"],
            item["description"],
            item["price"],
            item["image_url"],
            item["tags"],
            item["category"],
            item["specs"],
            item["selling_points"],
            item["policy"],
            item["paid_count"],
            item["review_count"],
            item["sales_volume"],
            item["positive_rate"],
            item["total_reviews"],
            item["shipping_time"],
            item["support_7day_return"],
            item["support_fast_refund"],
            item["payment_method"],
            item["store_name"],
            item["store_url"],
            item["top_reviews_json"],
            item["source_product_id"],
            item["data_source"],
            item["created_at"],
            item["updated_at"],
        )
        for item in products
    ]

    if not rows:
        return

    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO products (
              id, name, description, price, image_url, tags, category, specs,
              selling_points, policy,
              paid_count, review_count, sales_volume, positive_rate, total_reviews,
              shipping_time, support_7day_return, support_fast_refund, payment_method,
              store_name, store_url, top_reviews_json,
              source_product_id, data_source, created_at, updated_at
            ) VALUES %s
            ON CONFLICT (id) DO UPDATE SET
              name = EXCLUDED.name,
              description = EXCLUDED.description,
              price = EXCLUDED.price,
              image_url = EXCLUDED.image_url,
              tags = EXCLUDED.tags,
              category = EXCLUDED.category,
              specs = EXCLUDED.specs,
              selling_points = EXCLUDED.selling_points,
              policy = EXCLUDED.policy,
              paid_count = EXCLUDED.paid_count,
              review_count = EXCLUDED.review_count,
              sales_volume = EXCLUDED.sales_volume,
              positive_rate = EXCLUDED.positive_rate,
              total_reviews = EXCLUDED.total_reviews,
              shipping_time = EXCLUDED.shipping_time,
              support_7day_return = EXCLUDED.support_7day_return,
              support_fast_refund = EXCLUDED.support_fast_refund,
              payment_method = EXCLUDED.payment_method,
              store_name = EXCLUDED.store_name,
              store_url = EXCLUDED.store_url,
              top_reviews_json = EXCLUDED.top_reviews_json,
              source_product_id = EXCLUDED.source_product_id,
              data_source = EXCLUDED.data_source,
              updated_at = EXCLUDED.updated_at
            """,
            rows,
            template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
        )
    conn.commit()


def upsert_product_vectors(conn, products: list[dict]):
    rows = [build_vector_row(item) for item in products]

    if not rows:
        return

    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO product_vectors (id, product_id, chunk_index, chunk_text, metadata_json, source, updated_at)
            VALUES %s
            ON CONFLICT (product_id, chunk_index) DO UPDATE SET
              chunk_text = EXCLUDED.chunk_text,
              metadata_json = EXCLUDED.metadata_json,
              source = EXCLUDED.source,
              updated_at = EXCLUDED.updated_at
            """,
            rows,
            template="(%s, %s, %s, %s, %s, %s, %s)",
        )
    conn.commit()


def fetch_all_products(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
              id,
              name,
              category,
              price,
              specs,
              selling_points,
              policy,
              sales_volume,
              positive_rate,
              store_name,
              source_product_id,
              data_source,
              updated_at,
              created_at
            FROM products
            """
        )
        rows = cur.fetchall()

    products: list[dict] = []
    for row in rows:
        updated_at = row[12] or row[13] or datetime.now(timezone.utc)
        products.append(
            {
                "id": row[0],
                "name": row[1],
                "category": row[2],
                "price": row[3],
                "specs": row[4],
                "selling_points": row[5],
                "policy": row[6],
                "sales_volume": row[7],
                "positive_rate": row[8],
                "store_name": row[9],
                "source_product_id": row[10],
                "data_source": row[11],
                "updated_at": updated_at,
            }
        )

    return products


def fetch_vector_coverage(conn) -> tuple[int, int]:
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM products")
        product_total = int(cur.fetchone()[0])
        cur.execute("SELECT COUNT(DISTINCT product_id) FROM product_vectors")
        vector_total = int(cur.fetchone()[0])
    return product_total, vector_total


def fetch_source_counts(conn, source: str) -> tuple[int, int]:
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM products WHERE COALESCE(data_source, 'MANUAL') = %s", (source,))
        product_count = int(cur.fetchone()[0])
        cur.execute("SELECT COUNT(*) FROM product_vectors WHERE source = %s", (source,))
        vector_count = int(cur.fetchone()[0])
    return product_count, vector_count


def parse_args():
    parser = argparse.ArgumentParser(description="Import products from local CSV/XLSX and sync vectors")
    parser.add_argument("--input-csv", type=str, default="", help="Import products from local CSV file")
    parser.add_argument("--input-xlsx", type=str, default="", help="Import products from local XLSX file")
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT, help="Limit imported rows")
    parser.add_argument("--dry-run", action="store_true", help="Only parse and print summary, do not write DB")
    parser.add_argument("--replace-source", type=str, default="", help="Clear existing data by data_source before ingest, e.g. XLSX or CSV")
    parser.add_argument("--sync-all-vectors", action="store_true", help="Rebuild vectors for all products in database")
    return parser.parse_args()


def main():
    args = parse_args()
    limit = max(1, min(args.limit, MAX_LIMIT))

    input_file = ""
    if args.input_xlsx and args.input_xlsx.strip():
        input_file = args.input_xlsx.strip()
    elif args.input_csv and args.input_csv.strip():
        input_file = args.input_csv.strip()

    products: list[dict] = []
    if input_file:
        products = load_products(input_file, limit)
        products = deduplicate_products(products)
    elif not args.sync_all_vectors:
        raise ValueError("Please provide --input-xlsx or --input-csv, or use --sync-all-vectors")

    source_counter: dict[str, int] = {}
    for product in products:
        source_name = normalize_text(product.get("data_source") or "UNKNOWN")
        source_counter[source_name] = source_counter.get(source_name, 0) + 1

    summary = {
        "requested_limit": args.limit,
        "actual_limit": len(products),
        "source_breakdown": source_counter,
        "sample_product": {
            "id": str(products[0]["id"]) if products else "",
            "name": products[0]["name"] if products else "",
            "category": products[0]["category"] if products else "",
            "update_time": products[0]["updated_at"].isoformat() if products else "",
        } if products else {},
        "vector_rows": len(products),
    }

    if args.dry_run:
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
        return

    config = get_db_config()
    conn = psycopg2.connect(**config)
    try:
        ensure_schema(conn)
        if args.replace_source and args.replace_source.strip():
            source = normalize_text(args.replace_source).upper()
            with conn.cursor() as cur:
                cur.execute("DELETE FROM product_vectors WHERE source = %s", (source,))
                cur.execute("DELETE FROM cart_items WHERE product_id IN (SELECT id FROM products WHERE COALESCE(data_source, 'MANUAL') = %s)", (source,))
                cur.execute("DELETE FROM product_views WHERE product_id IN (SELECT id FROM products WHERE COALESCE(data_source, 'MANUAL') = %s)", (source,))
                cur.execute("DELETE FROM products WHERE COALESCE(data_source, 'MANUAL') = %s", (source,))
            conn.commit()

        if products:
            upsert_products(conn, products)
            upsert_product_vectors(conn, products)

        if args.sync_all_vectors:
            all_products = fetch_all_products(conn)
            upsert_product_vectors(conn, all_products)

        current_source = normalize_text(products[0].get("data_source") if products else "CSV") or "CSV"
        product_count, vector_count = fetch_source_counts(conn, current_source)
        total_products, total_vector_products = fetch_vector_coverage(conn)
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
        if products:
            print(f"Ingested {len(products)} {current_source} products and {len(products)} vector rows.")
            print(f"Database totals -> products({current_source})={product_count}, product_vectors({current_source})={vector_count}")
        print(f"Vector coverage -> products={total_products}, vectorized_products={total_vector_products}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
